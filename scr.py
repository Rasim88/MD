import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

# загружаем файл
wb = openpyxl.load_workbook('20230301.xlsx')

# выбираем имеющиеся листы таблицы
sheet1 = wb['Объем потребления(л)']
sheet2 = wb['Средняя цена']

# очищаем ячейки A1 и B1 (заголовок и комментарий)
sheet1['A1'].value = None
sheet1['B1'].value = None
sheet2['A1'].value = None
sheet2['B1'].value = None

# создаем новый лист для объединенных данных
merged_sheet = wb.create_sheet('Сводная таблица')

# объединяем заголовки
merged_sheet.append(['Дата', 'Код региона', 'Регион', 'АИ-95', 'АИ-92', 'АИ-98', 'ДТ', 'Метан', 'Пропан', 'КПГ', 'СУГ', 'Другие газы'])

# создаем словарь для хранения данных по регионам
data_dict1 = {}
data_dict2 = {}

# заполняем словарь данными из первого листа
for row in sheet1.iter_rows(min_row=2, values_only=True):
    region = row[1]
    if region not in data_dict1:
        data_dict1[region] = [row[0], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11]]
    else:
        print("Duplicate region code in sheet 1: ", region)

# заполняем словарь данными из второго листа
for row in sheet2.iter_rows(min_row=2, values_only=True):
    region = row[1]
    if region not in data_dict2:
        data_dict2[region] = [row[0], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11]]
    else:
        print("Duplicate region code in sheet 2: ", region)

# объединяем данные по регионам
for region, data_list1 in data_dict1.items():
    if region in data_dict2:
        data_list2 = data_dict2[region]
        merged_row = [data_list1[0], region, data_list1[1]] + data_list1[2:] + data_list2[2:]
        merged_sheet.append(merged_row)

# выбираем нужный лист таблицы
sheet3 = wb['Сводная таблица']

# удалить строки 1 и 2
sheet3.delete_rows(1, 2)

# устанавливаем порядок столбцов
column_order = ['A', 'B', 'C', 'D', 'M', 'E', 'N', 'F', 'O', 'G', 'P', 'H', 'Q', 'I', 'R', 'J', 'S', 'K', 'T', 'L', 'U']

# создаем словарь для хранения порядка столбцов
col_dict = {get_column_letter(idx): [] for idx in range(1, sheet3.max_column + 1)}
for col in sheet3.columns:
    col_dict[col[0].column_letter] = col

# очищаем лист
sheet3.delete_cols(1, sheet3.max_column)

# добавляем столбцы в нужном порядке
for idx, col in enumerate(column_order, start=1):
    # получаем номер столбца по буквенному обозначению
    col_num = column_index_from_string(col)
    # добавляем столбец на новую позицию
    new_col = col_dict[col]
    new_col[0].column = idx
    sheet3.insert_cols(idx)
    for row_num, cell in enumerate(new_col, start=1):
        sheet3.cell(row=row_num, column=idx, value=cell.value)


# присвоить имя ячейкам
cell_d1 = sheet3.cell(row=1, column=4, value='Н.п. АИ-92')
cell_e1 = sheet3.cell(row=1, column=5, value='Ср.цена АИ-92')
cell_f1 = sheet3.cell(row=1, column=6, value='Н.п. АИ-95')
cell_g1 = sheet3.cell(row=1, column=7, value='Ср.цена АИ-95')
cell_h1 = sheet3.cell(row=1, column=8, value='Н.п. АИ-98')
cell_i1 = sheet3.cell(row=1, column=9, value='Ср.цена АИ-98')
cell_j1 = sheet3.cell(row=1, column=10, value='Н.п. ДТ')
cell_k1 = sheet3.cell(row=1, column=11, value='Ср.цена ДТ')
cell_l1 = sheet3.cell(row=1, column=12, value='Н.п. Метан')
cell_m1 = sheet3.cell(row=1, column=13, value='Ср.цена Метан')
cell_n1 = sheet3.cell(row=1, column=14, value='Н.п. Пропан')
cell_o1 = sheet3.cell(row=1, column=15, value='Ср.цена Пропан')
cell_p1 = sheet3.cell(row=1, column=16, value='Н.п. КПГ')
cell_q1 = sheet3.cell(row=1, column=17, value='Ср.цена КПГ')
cell_r1 = sheet3.cell(row=1, column=18, value='Н.п. СУГ')
cell_s1 = sheet3.cell(row=1, column=19, value='Ср.цена СУГ')
cell_t1 = sheet3.cell(row=1, column=20, value='Н.п. Другие газы')
cell_u1 = sheet3.cell(row=1, column=21, value='Ср.цена Другие газы')

# сохраняем изменения в файл
wb.save('20230301_py.xlsx')